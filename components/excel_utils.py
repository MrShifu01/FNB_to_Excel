from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table,TableStyleInfo
import csv

def create_excel_from_csv(csv_file_path, excel_file_path):
    wb = Workbook()
    ws = wb.active
    # Open and read the CSV file, then write to Excel
    with open(csv_file_path, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    wb.save(excel_file_path)

def add_dataframe_to_excel(destination_file_path, df):
    # Load the destination Excel workbook
    
    wb = load_workbook(destination_file_path)
    ws = wb['Transactions']

    # If there's a table in the sheet, convert it to a range (remove the table)
    if ws.tables:
        table_name = list(ws.tables.keys())[0]
        del ws.tables[table_name]

    # Find the first empty row in column A
    first_empty_row = ws.max_row + 1

    # Write the DataFrame rows starting from the first empty row
    for index, row in df.iterrows():
        # start=1 because Excel columns are 1-indexed
        for col_num, cell_value in enumerate(row, start=1):
            ws.cell(row=first_empty_row + index, column=col_num, value=cell_value)

    return wb, ws

def deduplicate_rows(ws):
    # Determine the columns for which you want to check duplicates. E.g., "Date", "Description", and "Amount"
    columns_to_check = ["Account", "Date", "Year",
                        "Adj Month", "Description", "Amount"]
    column_indices = []

    for col_num, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1, values_only=True), start=1):
        if col_cells[0] in columns_to_check:
            column_indices.append(col_num)

    # Identify duplicate rows
    seen = set()
    rows_to_delete = []
    print(len(seen))
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        entry = tuple(row[i-1] for i in column_indices)

        if entry in seen:
            rows_to_delete.append(row_num)
        else:
            seen.add(entry)
    print(len(seen))

    # Delete duplicate rows (loop backwards to avoid modifying the rows while deleting)
    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num)
    return ws

def convert_range_to_table(ws):
    # Convert the range back to a table in openpyxl
    table = Table(displayName="Table1", ref=ws.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    return ws